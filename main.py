import openpyxl
import telebot
from telebot import types

API_TOKEN = '7635646554:AAHaH6Yi8ET705pT4QIaBjukWFJy1_VLKCs'
bot = telebot.TeleBot(API_TOKEN)

#переменная для хранения выбранного действия
selected_action = None

#преобразование в число
def safe_int(value):
    try:
        return float(value) if value else 0
    except ValueError:
        return 0

#для подсчета % выполнения дз
def calculate_homework_given_percentage(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    teacher_stats = {}

    for row in sheet.iter_rows(min_row=3, values_only=True):
        teacher = row[1]
        total_given = safe_int(row[3])
        total_planned = safe_int(row[6])

        if teacher is None or teacher == '':
            continue

        if teacher not in teacher_stats:
            teacher_stats[teacher] = {'given': 0, 'planned': 0}

        teacher_stats[teacher]['given'] += total_given
        teacher_stats[teacher]['planned'] += total_planned

    results = {}
    for teacher, stats in teacher_stats.items():
        if stats['planned'] > 0:
            percentage = round((stats['given'] / stats['planned']) * 100)
        else:
            percentage = 0
        results[teacher] = percentage

    return results

#подсчет % проверенных дз
def calculate_homework_checked_percentage(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    teacher_stats = {}

    for row in sheet.iter_rows(min_row=3, values_only=True):
        teacher = row[1]  #имя преподавателя
        total_checked = safe_int(row[5])
        total_received = safe_int(row[4])

        if teacher is None or teacher == '':
            continue

        if teacher not in teacher_stats:
            teacher_stats[teacher] = {'checked': 0, 'received': 0}

        teacher_stats[teacher]['checked'] += total_checked
        teacher_stats[teacher]['received'] += total_received

    results = {}
    for teacher, stats in teacher_stats.items():
        if stats['received'] > 0:
            percentage = round((stats['checked'] / stats['received']) * 100)
        else:
            percentage = 0
        results[teacher] = percentage

    return results

#отправка отчета в чат
def send_report(message, file_path):
    if selected_action == 'given':
        results = calculate_homework_given_percentage(file_path)
        report = "Отчет по выданным дз:\n"
    elif selected_action == 'checked':
        results = calculate_homework_checked_percentage(file_path)
        report = "Отчет по проверенным дз:\n"
    else:
        bot.send_message(message.chat.id, "Ошибка: не выбрано действие.")
        return

    if not results:
        report += "Нет данных по преподавателям или данные некорректны."
    else:
        for teacher, percentage in results.items():
            report += f"{teacher}: {percentage}% выполнено\n"

    bot.send_message(message.chat.id, report)

    #повтор кнопок
    show_action_buttons(message)

#кнопки для выбора действия
def show_action_buttons(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    btn1 = types.KeyboardButton("% выданных заданий")
    btn2 = types.KeyboardButton("% проверенных заданий")
    markup.add(btn1, btn2)

    bot.send_message(message.chat.id, "Следующее действие:", reply_markup=markup)

#команда /start
@bot.message_handler(commands=['start'])
def send_welcome(message):
    show_action_buttons(message)

#команда /stop
@bot.message_handler(commands=['stop'])
def stop_bot(message):
    bot.send_message(message.chat.id, "Бот остановлен. Для запуска используй /start.")


#обработка действий
@bot.message_handler(func=lambda message: message.text in ["% выданных заданий", "% проверенных заданий"])
def handle_action_selection(message):
    global selected_action

    if message.text == "% выданных заданий":
        selected_action = 'given'
    elif message.text == "% проверенных заданий":
        selected_action = 'checked'
    bot.send_message(message.chat.id, "Йоу! Отправь мне Excel файл с данными.", reply_markup=types.ReplyKeyboardRemove())

#для получения файла
@bot.message_handler(content_types=['document'])
def handle_document(message):
    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)

    # сохранение файла
    with open("homework_data.xlsx", "wb") as f:
        f.write(downloaded_file)
    #отчёт
    send_report(message, "homework_data.xlsx")

bot.polling()

